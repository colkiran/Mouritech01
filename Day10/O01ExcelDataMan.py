
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

wb.active = wb['Players']

ws = wb.active

# A6:E11    =   A - col1, row6, E - col5 row11
# min-row = 6, max_row = 11, min_col = 1, max_col = 5

# for row in ws.iter_rows(min_row=6, min_col=1, max_row=11, max_col=5):

# for row in ws.iter_rows(min_row=6, min_col=1, max_row=11):
#     for cell in row:
#         if cell.value == 'Lara':
#             cell.value = 'brain lara'.upper()

for row in ws.iter_rows(min_row=7, min_col=4, max_col=4, max_row=11):
    for cell in row:
        cell.value += 10

wb.save("FirstExcel.xlsx")

# increase the mathces by 10 no


