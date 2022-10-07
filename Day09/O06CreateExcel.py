
from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

ws = wb.active
ws.title = "MySheet"

ws["C5"] = "Hello World"
ws['C6'] = 10000
ws['C7'] = datetime.now()

wb.save("FirstExcel.xlsx")
